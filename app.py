# relevant sources:
# - dash datatable example with callback: https://dash.plotly.com/datatable/editable

# -*- coding: utf-8 -*-
import base64
import io
import dash
import dash_core_components as dcc
import dash_html_components as html
# import dash_bootstrap_components as dbc
import plotly.graph_objs as go
import pandas as pd
import pathlib
import numpy as np
import dash_table
import os
from urllib.parse import quote as urlquote
from dash.dependencies import Input, Output, State
from scipy import stats
from openpyxl import load_workbook
from ui_elements import create_app_layout
import pdfplumber
import json
import tabula
from selenium import webdriver
from webdriver_manager.chrome import ChromeDriverManager
import subprocess
from init import UPLOAD_DIRECTORY, SCRAPED_DIRECTORY, UPLOAD_PARSED_DIRECTORY
import pickle


def get_webdriver_path(driver_name):
    path = subprocess.check_output(f'which {driver_name}', shell=True)
    path = path.decode('utf-8').replace('\n', '')
    return path

with open('data/onedrive_dict.pkl', 'rb') as file:
    onedrive_dict = pickle.load(file)

onedrive_dict_ui = {}
for child in onedrive_dict['l0']:
    l_l1 = onedrive_dict['l1'][child.id]
    if len(l_l1) > 0:
        for grandchild in l_l1:
            onedrive_dict_ui[child.name + '/' + grandchild.name] = grandchild.web_url
    else:
        onedrive_dict_ui[child.name] = child.web_url

def save_file(name, content):
    """Decode and store a file uploaded with Plotly Dash."""
    data = content.encode("utf8").split(b";base64,")[1]
    with open(os.path.join(UPLOAD_DIRECTORY, name), "wb") as fp:
        fp.write(base64.decodebytes(data))


def file_download_link(filename):
    """Create a Plotly Dash 'A' element that downloads a file from the app."""
    location = "/download/{}".format(urlquote(filename))
    return html.A(filename, href=location)


def uploaded_files(dir=UPLOAD_DIRECTORY, ext=None):
    """List the files in the upload directory."""
    files = []
    for filename in os.listdir(dir):
        path = os.path.join(dir, filename)
        if os.path.isfile(path):
            files.append(filename)
    if ext is not None:
        files = [file for file in files if ext in file]

    return files


def parse_excel_table(ws, data_boundary):
    """
    parse excel table with given boundary from excel sheets
    :param ws:
    :param data_boundary:
    :return:
    """

    # parse the data within the ref boundary
    data = ws[data_boundary]
    # extract the data
    # the inner list comprehension gets the values for each cell in the table
    content = [[cell.value for cell in ent] for ent in data]

    header = content[0]

    # the contents ... excluding the header
    rest = content[1:]

    # create dataframe with the column names
    # and pair table name with dataframe
    df = pd.DataFrame(rest, columns=header)

    return df


def extract_excel_tables(file_path):
    l_tables_html = []  # list of html tables
    filename, file_extension = os.path.splitext(file_path)
    final_data = load_workbook(file_path)
    mapping = {}
    for sheet_name in final_data.sheetnames:
        ws = final_data[sheet_name]
        for entry, data_boundary in ws.tables.items():
            mapping[entry] = parse_excel_table(ws, data_boundary)
            # convert to dict
            data = mapping[entry].to_dict(orient='records')
            columns = [{'name': str(i), 'id': str(i)} for i in mapping[entry].columns]
            html_div = html.Div([
                html.H6(f"{filename}_{sheet_name}_{entry}"),
                dash_table.DataTable(
                    id='table-editing-simple',
                    editable=True,
                    data=data,
                    columns=columns)
            ])
            l_tables_html = l_tables_html + [html_div]

    return l_tables_html


def parse_pdf_table_to_df(table):
    row_num = 0
    for row in table['data']:
        l_items = []
        for item in row:
            if item is not None:
                l_items = l_items + [item['text'].replace(u'\xa0', u' ').replace(u'\n', '')]
            else:
                l_items = l_items + ['']
        if row_num == 0:
            df = pd.DataFrame(columns=l_items)
        else:
            df.loc[row_num] = l_items
        row_num = row_num + 1

    return df


def extract_pdf_tables(file_path):
    l_tables = tabula.read_pdf(file_path,
                               # stream=True,
                               # lattice=True,
                               pages='all',
                               multiple_tables=True,
                               output_format='json')
    l_tables = [table for table in l_tables if len(table['data']) > 0]
    l_table_html = []
    table_num = 0
    for table in l_tables:
        try:
            df = parse_pdf_table_to_df(table)
            if np.unique(df.values).all() in ['']:
                print(f'all empty values in table number {table_num}')
                table_num = table_num + 1
                continue
            data = df.to_dict(orient='records')
            columns = [{'name': str(i), 'id': str(i)} for i in df.columns]
            html_div = html.Div([
                html.H6(f"{file_path}_{table_num}"),
                dash_table.DataTable(
                    id='table-editing-simple',
                    editable=True,
                    data=data,
                    columns=columns)
            ])
            l_table_html = l_table_html + [html_div]
        except Exception as e:
            print('error in extract_pdf_tables()')
            print(e)
        table_num = table_num + 1

    return l_table_html


def extract_pdf_text(file_path):
    filename, file_extension = os.path.splitext(file_path)
    l_text_html = []
    l_table_html = []
    df_text = pd.DataFrame()
    pdf_doc = pdfplumber.open(file_path)
    page_num = 0
    for page in pdf_doc.pages:
        df = pd.DataFrame()
        text = page.extract_text()
        df['page_num'] = [page_num]
        df['text'] = [text]
        df_text = df_text.append(df)
        data = df_text.to_dict(orient='records')
        columns = [{'name': str(i), 'id': str(i)} for i in df_text.columns]
        html_div = html.Div([
            html.P(text)
        ])
        l_text_html = l_text_html + [html_div]
        page_num = page_num + 1

    return l_text_html  # , l_table_html


group_colors = {"control": "light blue", "reference": "red"}

app = dash.Dash(
    __name__, meta_tags=[{"name": "viewport", "content": "width=device-width"}]
)
server = app.server

PATH = pathlib.Path(__file__).parent
DATA_PATH = PATH.joinpath("data").resolve()
# default_study_data = pd.read_csv(DATA_PATH.joinpath("study.csv"))
default_study_data = load_workbook(DATA_PATH.joinpath("sample_excel.xlsx"))

# App Layout
app.layout = create_app_layout(app)

# update the list of onedrive files available
@app.callback(
    [Output("onedrive-dropdown", "options"),
     Output("onedrive-dropdown", "value"),
     Output("onedrive-source-dropdown", "options"),
     Output("onedrive-source-dropdown", "value")],
    [Input('tabs', 'value')],
)
def update_cloud_files(tab):
    # print(f'tab: {tab}')
    if tab == 'tab_onedrive':
        # print('inside update_cloud_files()/tab_onedrive')

        l_onedrive_ui = []
        for key in list(onedrive_dict_ui.keys()):
            l_onedrive_ui = l_onedrive_ui + [{'label': key, 'value': key}]
        # print(l_onedrive_ui)
        return l_onedrive_ui, l_onedrive_ui[0]['value'], l_onedrive_ui, l_onedrive_ui[0]['value']
    else:
        return [], [], [], []

# update the list of raw files uploaded
@app.callback(
    [Output("file-dropdown", "options"), Output("file-dropdown", "value")],
    [Input("upload-data", "filename"), Input("upload-data", "contents")],
)
def update_file_dropdown(uploaded_filenames, uploaded_file_contents):
    """Save uploaded files and regenerate the file list."""
    if uploaded_filenames is not None and uploaded_file_contents is not None:
        for name, data in zip(uploaded_filenames, uploaded_file_contents):
            save_file(name, data)
    files = uploaded_files(UPLOAD_DIRECTORY)
    if len(files) == 0:
        return [html.Li("No files yet!")]
    else:
        l_files_dict = []
        for file in files:
            l_files_dict = l_files_dict + [{'label': file, 'value': file}]

        return l_files_dict, files[0]  # dropdown_ui  # [html.Li(file_download_link(filename)) for filename in files]

# update the list of tables extracted from raw files
@app.callback(
    [Output("table-dropdown", "options"), Output("table-dropdown", "value")],
    [Input("file-dropdown", "value")],  # Input("upload-data", "contents")
)
def update_table_dropdown(filenames):  # uploaded_file_contents
    # """Save uploaded files and regenerate the file list."""
    # if uploaded_filenames is not None:  #  and uploaded_file_contents is not None
    #     for name, data in zip(uploaded_filenames, uploaded_file_contents):
    #         save_file(name, data)
    files = uploaded_files(UPLOAD_PARSED_DIRECTORY, ext='.csv')
    if len(files) == 0:
        return [], []  # [html.Li("No files yet!")]
    else:
        l_files_dict = []
        for file in files:
            l_files_dict = l_files_dict + [{'label': file, 'value': file}]

        return l_files_dict, files[0]  # dropdown_ui  # [html.Li(file_download_link(filename)) for filename in files]


# update the list of columns extracted from selected tables
@app.callback(
    [Output("column-dropdown", "options"), Output("column-dropdown", "value")],
    [Input("table-dropdown", "value")],  # Input("upload-data", "contents")
)
def update_column_dropdown(filenames):  # uploaded_file_contents
    """Save uploaded files and regenerate the file list."""
    l_cols_dict = []
    l_cols = []
    if len(filenames) == 0:
        filenames = [os.listdir(UPLOAD_PARSED_DIRECTORY)[-1]]
    if not isinstance(filenames, list):
        filenames = [filenames]

    # print(filenames)
    l_files_dict = []
    for filename in filenames:
        file = os.path.join(UPLOAD_PARSED_DIRECTORY, filename)
        df = pd.read_csv(file)
        cols = df.columns
        l_cols = l_cols + cols.to_list()
        l_cols_dict = l_cols_dict + [{'label': col, 'value': col} for col in cols]

        return l_cols_dict, l_cols


UPLOAD_MODE = True
@app.callback(
    [
        Output('tables-tab', 'children'),
        Output('text-tab', 'children')
        # Output('table-editing-simple', 'data'),
        # Output('table-editing-simple', 'columns')
    ],
    [Input("table-dropdown", "value"), Input("column-dropdown", "value")],  # Input("upload-data", "filename"),
    # [State("upload-data", "contents")],  # State("error", "data")
)
def display_output(uploaded_filenames, table_columns):  # uploaded_contents
    l_tables_html = []
    l_text_html = []
    if uploaded_filenames and UPLOAD_MODE:
        for filename in uploaded_filenames:
            file_path = os.path.join(UPLOAD_PARSED_DIRECTORY, filename)
            # print(file_path)
            filename_noext, file_extension = os.path.splitext(file_path)
            # Try reading uploaded file
            if file_extension in ['.xlsx']:
                try:
                    l_tables_html = l_tables_html + \
                                    extract_excel_tables(file_path)
                # Data is invalid
                except Exception as e:
                    error_message = html.Div(
                        className="alert",
                        children=["unable to parse spreadsheet"],
                    )
                    error_status = True
            elif file_extension in ['.pdf']:
                error_message = html.Div(
                    className="alert",
                    children=["processing pdf file"],
                )
                try:
                    l_tables_html = l_tables_html + extract_pdf_tables(file_path)
                # Data is invalid
                except Exception as e:
                    print('unable to run extract_pdf_tables()')
                    error_message = html.Div(
                        className="alert",
                        children=[f"unable to parse {filename}"],
                    )
                    error_status = True
                try:
                    l_text_html = l_text_html + extract_pdf_text(file_path)
                except Exception as e:
                    print('unable to run extract_pdf_tables()')
            elif file_extension in ['.csv']:
                df = pd.read_csv(file_path)
                cols_keep = [col for col in df.columns.tolist() if col in table_columns]
                df = df[cols_keep]
                data = df.to_dict(orient='records')
                columns = [{'name': str(i), 'id': str(i)} for i in df.columns]
                html_div = html.Div([
                    html.H6(f"{filename.replace('.csv', '')}"),
                    dash_table.DataTable(
                        id='table-editing-simple',
                        editable=True,
                        data=data,
                        columns=columns)
                ])
                l_tables_html = l_tables_html + [html_div]
    else:    # if no uploaded files
        l_files = os.listdir(UPLOAD_DIRECTORY)
        l_files = [file for file in l_files if '.csv' in file]
        dict_df = {}
        for file in l_files:
            df = pd.read_csv(os.path.join(UPLOAD_DIRECTORY, file))
            dict_df[file] = df
        # subset over selected columns
        df = dict_df['swiss_re_financials.csv']
        data = df.to_dict(orient='records')
        columns = [{'name': str(i), 'id': str(i)} for i in df.columns]
        fig = {
            'data': [{
                'type': 'parcoords',
                'dimensions': [{
                    'label': col['name'],
                    'values': df[col['id']]
                } for col in columns]
            }]
        }

        # data from uploaded content
        l_tables_html = l_tables_html + \
                        [html.Div(dash_table.DataTable(
                            id='table-editing-simple',
                            editable=True,
                            data=data,
                            columns=columns
                        ))]

    tables_tab = [html.Div(l_tables_html)]
    text_tab = [html.Div(l_text_html)]

    return tables_tab, text_tab

# update images tab
@app.callback(
    [Output('images-tab', 'children')],
    [Input("web_search_button", "n_clicks"),
     Input("web-search-sources", "value")]
)
def scrape_web(n_clicks, urls):
    if n_clicks > 0:
        print('inside scrape_web(): callback for search button')
        # driver_path = get_webdriver_path(driver_name='geckodriver')
        driver_path = ChromeDriverManager().install()
        driver = webdriver.Chrome(executable_path=driver_path)  # Chrome()
        driver.get(url=urls[0])
        print('loaded url')
        file_name = f'{SCRAPED_DIRECTORY}/scb_test.html'

        # wrtie html file
        html_file = open(file_name, "w")
        html_file.write(driver.page_source)  # scraper.
        html_file.close()
        driver.quit()
        print('finished scrape_web(): callback for search button')

    return []


# # update iframe tab
# @app.callback(
#     [Output('iframe_tab', 'children')],
#     [Input("onedrive-dropdown", "value")]
# )
# def update_iframe(file_name):
#     if len(file_name) > 0:
#         # print(file_name)
#         # print(onedrive_dict_ui[file_name])
#         iframe_url = [html.Iframe(
#             src=onedrive_dict_ui[file_name],
#             style={"height": "1067px", "width": "100%"})]
#         return iframe_url
#     else:
#         return [html.Iframe(
#             src="https://1drv.ms/f/s!ABwr6VEMK7BskwM",
#             style={"height": "1067px", "width": "100%"})]



if __name__ == "__main__":
    app.run_server(host='localhost', port=8500, debug=True)





