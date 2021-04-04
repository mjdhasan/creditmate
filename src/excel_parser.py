# https://samukweku.github.io/data-wrangling-blog/spreadsheet/python/pandas/openpyxl/2020/05/19/Access-Tables-In-Excel.html
from openpyxl import load_workbook
import pandas as pd
import os
from xls2xlsx import XLS2XLSX
from init import DIR_DUMPS

# DIR_PATH = DIR_DUMPS
DIR_PATH = os.path.join(DIR_DUMPS, '016-TextFiles')
l_files = os.listdir(DIR_PATH)
filename = l_files[4]
file_path = os.path.join(DIR_PATH, filename)
filename, file_extension = os.path.splitext(file_path)

## convert xls file to xlsx
# xls_file = XLS2XLSX(file_path)
# file_path = f"{file_path.replace('.xls', '')}.xlsx"
# xls_file.to_xlsx(file_path)

# read using pandas
df = pd.read_excel(file_path)
df.columns
df.to_html(f"{file_path.replace('.xlsx', '')}.html")


#read file
wb = load_workbook(file_path)
dir(wb)
wb.get_sheet_names()
# wb['15']
sheet_name = [sheet for sheet in wb.get_sheet_names() if 'economic_indicator' in sheet][0]

#access specific sheet
ws = wb[wb.sheetnames[3]]  # wb["Tables"]
dir(ws)
ws.merged_cell_ranges[0]
ws.merged_cell_ranges[0].coord
(ws.merged_cell_ranges[0].min_col, ws.merged_cell_ranges[0].max_col)
ws.merged_cell_ranges[0].shrink()
ws.merged_cell_ranges[0].to_tree()
[cell for cell in ws.merged_cell_ranges[0].cells]
ws.title
ws.tables
[value for value in ws.values]
ws.sheet_format
ws.show_gridlines()
# ws._tables
type(ws.tables)
len(ws.tables)
l_tables = [table for table in ws.tables]
dir(l_tables[0])
l_tables[0].title()

{key : value for key, value in ws.tables.items()}

mapping = {}
for entry, data_boundary in ws.tables.items():
    # parse the data within the ref boundary
    data = ws[data_boundary]
    # extract the data
    # the inner list comprehension gets the values for each cell in the table
    content = [[cell.value for cell in ent] for ent in data]

    header = content[0]

    # the contents ... excluding the header
    rest = content[1:]

    # create dataframe with the column names
    # and pair table name with dataframe
    df = pd.DataFrame(rest, columns=header)
    mapping[entry] = df

mapping.keys()
mapping['dSupplier']

ws.max_row

ws[f'E23'].value

l_rows = list(ws.rows)
len(l_rows)
for row in l_rows[0:5]:
    # print(row.value)
    l_cells = [cell for cell in row]
    for cell in l_cells:
        print(cell.value)
